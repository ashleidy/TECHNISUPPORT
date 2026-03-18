
import sys


# Generación de archivos Excel
import openpyxl
from openpyxl.styles import  Alignment


# Generación de archivos PDF
from reportlab.lib import colors
from reportlab.lib.enums import TA_JUSTIFY

from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

# PyQt5 para la interfaz gráfica
from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QMessageBox,
    QHeaderView,
    QAbstractScrollArea,
    QFileDialog   
)
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import  TA_JUSTIFY
from class_DatabaseManager import DatabaseManager

class Tab4(QWidget):
    def __init__(self,connection_data):
        super().__init__()
        self.db = DatabaseManager(**connection_data)  # Pasar los datos de conexión
        self.initUI()
        self.load_data()

    def initUI(self):
        # Layout principal
        
        self.resize(900, 500)
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Campo de búsqueda
        self.search_label = QLabel("Buscar por Nombre o Cédula/NIT:")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Ingrese nombre o cédula/NIT...")
        self.search_button = QPushButton("Buscar")
        self.search_button.clicked.connect(self.search_data)

        search_layout = QHBoxLayout()
        search_layout.addWidget(self.search_label)
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_button)
        self.layout.addLayout(search_layout)

        # Tabla para mostrar datos
        self.table = QTableWidget()
        self.table.setFont(QFont("Arial", 10))
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #f8f9fa; /* Fondo claro */
                border: 1px solid #ddd; /* Borde suave */
                border-radius: 5px;
                gridline-color: #ddd; /* Color de las líneas de la cuadrícula */
            }
            QHeaderView::section {
                background-color: #2E86C1; /* Azul oscuro para los encabezados */
                color: white;
                padding: 10px;
                font-size: 14px;
                font-weight: bold;
                border: none;
            }
            QTableWidget::item {
                padding: 8px;
                font-size: 14px;
                border-bottom: 1px solid #ddd; /* Línea divisoria entre filas */
            }
            QTableWidget::item:selected {
                background-color: #2E86C1; /* Azul oscuro para elementos seleccionados */
                color: white;
            }
        """)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)  # Ajustar columnas al contenido
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)  # Ajustar el tamaño de la tabla al contenido
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # Habilitar barra de desplazamiento horizontal
        self.layout.addWidget(self.table)

        # Botones
        self.excel_button = QPushButton("Generar Excel")
        self.excel_button.setEnabled(False)
        self.excel_button.clicked.connect(self.generate_excel)

        self.pdf_button = QPushButton("Generar PDF")
        self.pdf_button.setEnabled(False)
        self.pdf_button.clicked.connect(self.generate_pdf)
        
        self.refrescar_button = QPushButton("Refrescar")
        self.refrescar_button.setEnabled(True)
        self.refrescar_button.clicked.connect(self.refresh_data)
        
        self.salir_button = QPushButton("Salir")
        self.salir_button.setEnabled(True)
        self.salir_button.clicked.connect(self.salir)

        # Estilos modernos para los botones (tonos de azul)
        button_style = """
            QPushButton {
                background-color: #2E86C1; /* Azul oscuro */
                border: none;
                color: white;
                padding: 10px 20px;
                text-align: center;
                text-decoration: none;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
                margin: 5px;
            }
            QPushButton:hover {
                background-color: #1B4F72; /* Azul más oscuro al pasar el mouse */
            }
            QPushButton:pressed {
                background-color: #154360; /* Azul aún más oscuro al presionar */
            }
            QPushButton:disabled {
                background-color: #85C1E9; /* Azul claro para botones deshabilitados */
            }
        """
        self.search_button.setStyleSheet(button_style)
        self.excel_button.setStyleSheet(button_style)
        self.pdf_button.setStyleSheet(button_style)
        self.refrescar_button.setStyleSheet(button_style)
        self.salir_button.setStyleSheet(button_style)

        # Layout para centrar los botones
        button_layout = QHBoxLayout()
        button_layout.addStretch(1)  # Espacio a la izquierda
        button_layout.addWidget(self.excel_button)
        button_layout.addWidget(self.pdf_button)
        button_layout.addWidget(self.refrescar_button)
        button_layout.addWidget(self.salir_button)
        button_layout.addStretch(1)  # Espacio a la derecha
        self.layout.addLayout(button_layout)

        # Conectar señal de selección de la tabla
        self.table.itemSelectionChanged.connect(self.update_button_states)
        
    def refresh_data(self):
            
        results, headers = self.db.load_data()  # Llama a la función load_data para obtener los datos

        if results and headers:
            # Limpiar la tabla antes de cargar nuevos datos
            self.table.clear()

            # Configurar las columnas y encabezados
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            self.table.setRowCount(len(results))

            # Llenar la tabla con los nuevos datos
            for row_idx, row_data in enumerate(results):
                for col_idx, cell_data in enumerate(row_data):
                    self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(cell_data) if cell_data is not None else ""))
        
        
    def load_data(self):
           
        results, headers = self.db.load_data()  # Llama a la función load_data

        if results and headers:
            # Mostrar los datos en la tabla
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            self.table.setRowCount(len(results))

            for row_idx, row_data in enumerate(results):
                for col_idx, cell_data in enumerate(row_data):
                    self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(cell_data) if cell_data is not None else ""))
        

    def search_data(self):
   
        search_text = self.search_input.text()
        results, headers = self.db.search_data(search_text)

        if results and headers:
            # Mostrar los datos en la tabla
            self.table.setColumnCount(len(headers))
            self.table.setHorizontalHeaderLabels(headers)
            self.table.setRowCount(len(results))

            for row_idx, row_data in enumerate(results):
                for col_idx, cell_data in enumerate(row_data):
                    self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(cell_data) if cell_data is not None else ""))
        else:
            QMessageBox.warning(self, "Error", "No se encontraron resultados o hubo un error en la consulta.")
            
    def update_button_states(self):
        # Habilitar los botones si hay una selección
        self.excel_button.setEnabled(bool(self.table.selectedItems()))
        self.pdf_button.setEnabled(bool(self.table.selectedItems()))
        


    def generate_pdf(self):
        try:
            # Obtener la fila seleccionada
            selected_items = self.table.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "Selección requerida", "Por favor, seleccione un registro de la tabla.")
                return

            row = selected_items[0].row()

            # Obtener datos de la fila seleccionada
            def get_cell_value(col):
                item = self.table.item(row, col)
                return item.text() if item else "N/A"

            # Abrir un diálogo para que el usuario elija la ubicación y el nombre del archivo
            options = QFileDialog.Options()
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar archivo PDF",  # Título del diálogo
                "",  # Directorio inicial (vacío para el predeterminado)
                "PDF Files (*.pdf);;All Files (*)",  # Filtros de archivo
                options=options,
            )

            # Si el usuario no selecciona una ubicación, salir de la función
            if not file_name:
                QMessageBox.warning(self, "Advertencia", "No se seleccionó una ubicación para guardar el archivo.")
                return

            # Asegurarse de que el archivo tenga la extensión .pdf
            if not file_name.endswith(".pdf"):
                file_name += ".pdf"

            # Crear el archivo PDF
            custom_size = (800, 792)  # 800 puntos de ancho x 792 puntos de alto (misma altura que letter)

            doc = SimpleDocTemplate(
                file_name,
                pagesize=custom_size,  # Usar el tamaño personalizado
                leftMargin=0.5*inch,
                rightMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch
            )

            # Crear lista de elementos para el PDF
            elements = []
            styles = getSampleStyleSheet()

            # Estilos personalizados
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=15,
                spaceAfter=30,
                alignment=1  # Centro
            )

            normal_style = ParagraphStyle(
                name='Justified',
                parent=styles['Normal'],
                alignment=TA_JUSTIFY,
                fontSize=14,
                leading=16,
                keepWithNext=True  # Asegura que los párrafos se alineen bien
            )
            
             # Título datos del equipo
            elements.append(Paragraph("DATOS DEL USUARIO:", title_style))

            # Información del cliente
            client_data = [
                ['Cliente:', get_cell_value(1)],
                ['Cedula / NIT:', get_cell_value(2)],
                ['Correo:', get_cell_value(3)],
                ['Teléfono:', get_cell_value(4)],
                ['Fecha de ingreso:', get_cell_value(6)],
                ['Fecha de entrega:', get_cell_value(7)],
            ]
            
         

            client_table = Table(client_data, colWidths=[2.5*inch, 3.4*inch])
            client_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('PADDING', (0, 0), (-1, -1), 3),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            client_table.hAlign = 'LEFT'  # Esta línea alinea la tabla a la izquierda en la página


            elements.append(client_table)
            elements.append(Spacer(1, 20))

            # Título datos del equipo
            elements.append(Paragraph("DATOS DEL EQUIPO:", title_style))

          # Información básica del equipo
            basic_data = [
                ['Tipo:', get_cell_value(8), 'Marca:', get_cell_value(9)],
                ['Modelo:', get_cell_value(10), 'N° Serie:', get_cell_value(11)]
            ]
            
            basic_table = Table(basic_data, colWidths=[1*inch, 1.5*inch, 1*inch, 1.5*inch])
            basic_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('BACKGROUND', (2, 0), (2, -1), colors.lightgrey),
                ('PADDING', (0, 0), (-1, -1), 3),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ]))
            
            basic_table.hAlign = 'LEFT'
            elements.append(basic_table)
            elements.append(Spacer(1, 10))

            # Main Board
            
           
            valores_celdas = [get_cell_value(12), get_cell_value(13), get_cell_value(14)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                main_board_data  = [['Main Board:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'N° Serie:', valores_celdas[2] ]]
                
           
                main_board_table = Table(main_board_data, colWidths=[1.2*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch])
                main_board_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                  
                    
                ]))
                elements.append(main_board_table)
                main_board_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Wifi
            valores_celdas = [get_cell_value(15), get_cell_value(16), get_cell_value(17)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                wifi_data  = [['WiFi:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'N° Serie:', valores_celdas[2] ]]
                
             
            

                wifi_table = Table(wifi_data, colWidths=[1.2*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch])
                wifi_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(wifi_table)
                wifi_table .hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Teclado
            valores_celdas = [get_cell_value(18), get_cell_value(19), get_cell_value(20)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                teclado_data  = [['Teclado:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'N° Serie:', valores_celdas[2] ]]
                
            

                teclado_table = Table(teclado_data, colWidths=[1.2*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch])
                teclado_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(teclado_table)
                teclado_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # CPU1
             # Verificar si todos los valores son "N/A" antes de agregar la tabla
             
            valores_celdas = [get_cell_value(21), get_cell_value(22), get_cell_value(23)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                cpu1_data = [['CPU1:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'N° Serie:', valores_celdas[2] ]]
                 

                cpu1_table = Table(cpu1_data, colWidths=[1.2*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch])
                cpu1_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(cpu1_table)
                cpu1_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # CPU2
            
            valores_celdas = [get_cell_value(24), get_cell_value(25), get_cell_value(26)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                cpu2_data = [['CPU2:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'N° Serie:', valores_celdas[2] ]]
                
          
                cpu2_table = Table(cpu2_data, colWidths=[1.2*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch])
                cpu2_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(cpu2_table)
                cpu2_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # PSU
            
            valores_celdas = [get_cell_value(27), get_cell_value(28), get_cell_value(29)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                psu_data = [['PSU:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2] ]]
                
            
            
                psu_table = Table(psu_data, colWidths=[1.2*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch, 1*inch, 1.7*inch])
                psu_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(psu_table)
                psu_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Pantalla
            valores_celdas = [get_cell_value(30), get_cell_value(31), get_cell_value(32), get_cell_value(33)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                pantalla_data = [['Pantalla:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
            
     

                pantalla_table = Table(pantalla_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch  ])
                pantalla_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(pantalla_table)
                pantalla_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Batería
            valores_celdas = [get_cell_value(34), get_cell_value(35), get_cell_value(37), get_cell_value(36)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                bateria_data = [['Batería:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
 

                bateria_table = Table(bateria_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                bateria_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(bateria_table)
                bateria_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # RAM1
            valores_celdas = [get_cell_value(38), get_cell_value(39), get_cell_value(40), get_cell_value(41)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                ram1_data = [['RAM1:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
             
           
                ram1_table = Table(ram1_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                ram1_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(ram1_table)
                ram1_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # RAM2
            valores_celdas = [get_cell_value(42), get_cell_value(43), get_cell_value(44), get_cell_value(45)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                ram2_data = [['RAM2:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
        
                ram2_table = Table(ram2_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                ram2_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(ram2_table)
                ram2_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # RAM3
            valores_celdas = [get_cell_value(46), get_cell_value(47), get_cell_value(48), get_cell_value(49)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                ram3_data = [['RAM3:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                

                ram3_table = Table(ram3_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                ram3_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(ram3_table)
                ram3_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # RAM4
            valores_celdas = [get_cell_value(50), get_cell_value(51), get_cell_value(52), get_cell_value(53)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                ram4_data = [['RAM4:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
     
                ram4_table = Table(ram4_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                ram4_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(ram4_table)
                ram4_table .hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 1
            valores_celdas = [get_cell_value(54), get_cell_value(55), get_cell_value(56), get_cell_value(57)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento1_data  = [['Unidad A1:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
             
          

                almacenamiento1_table = Table(almacenamiento1_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento1_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento1_table)
                almacenamiento1_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 2
            valores_celdas = [get_cell_value(58), get_cell_value(59), get_cell_value(60), get_cell_value(61)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento2_data  = [['Unidad A2:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
       
            

                almacenamiento2_table = Table(almacenamiento2_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento2_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento2_table)
                almacenamiento2_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 3
            valores_celdas = [get_cell_value(62), get_cell_value(63), get_cell_value(64), get_cell_value(65)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento3_data  = [['Unidad A3:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
          

                almacenamiento3_table = Table(almacenamiento3_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento3_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento3_table)
                almacenamiento3_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 4
           
            valores_celdas = [get_cell_value(66), get_cell_value(67), get_cell_value(68), get_cell_value(69)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento4_data  = [['Unidad A4:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
         

                almacenamiento4_table = Table(almacenamiento4_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento4_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento4_table)
                almacenamiento4_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 5
            valores_celdas = [get_cell_value(70), get_cell_value(71), get_cell_value(72), get_cell_value(73)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento5_data  = [['Unidad A5:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
                
                almacenamiento5_table = Table(almacenamiento5_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento5_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1),8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento5_table)
                almacenamiento5_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 6
            valores_celdas = [get_cell_value(74), get_cell_value(75), get_cell_value(76), get_cell_value(77)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento6_data  = [['Unidad A6:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
           
               

                almacenamiento6_table = Table(almacenamiento6_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento6_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento6_table)
                almacenamiento6_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 7
            
            valores_celdas = [get_cell_value(78), get_cell_value(79), get_cell_value(80), get_cell_value(81)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento7_data  = [['Unidad A6:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
           

                almacenamiento7_table = Table(almacenamiento7_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento7_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento7_table)
                almacenamiento7_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad de almacenamiento 8
        
            valores_celdas = [get_cell_value(82), get_cell_value(83), get_cell_value(84), get_cell_value(85)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                almacenamiento8_data = [['Unidad A8:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
            
    
                almacenamiento8_table = Table(almacenamiento8_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                almacenamiento8_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(almacenamiento8_table)
                almacenamiento8_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # GPU1
          
            # Obtener los valores de las celdas
            valores_celdas = [get_cell_value(86), get_cell_value(87), get_cell_value(88), get_cell_value(89)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                gpu1_data = [['GPU1:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
      
    

                gpu1_table = Table(gpu1_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                gpu1_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(gpu1_table)
                gpu1_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # GPU2
            valores_celdas = [get_cell_value(90), get_cell_value(91), get_cell_value(92), get_cell_value(93)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                gpu2_data = [['GPU2:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
             
         
                gpu2_table = Table(gpu2_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                gpu2_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(gpu2_table)
                gpu2_table .hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad DVD1
            valores_celdas = [get_cell_value(94), get_cell_value(95), get_cell_value(96), get_cell_value(97)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                dvd1_data  = [['Unidad DVD1:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
           

                dvd1_table = Table(dvd1_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                dvd1_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(dvd1_table)
                dvd1_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Unidad DVD2
            valores_celdas = [get_cell_value(98), get_cell_value(99), get_cell_value(100), get_cell_value(101)]

            # Verificar si todos los valores son 'N/A'
            if not all(valor == 'N/A' for valor in valores_celdas):
                dvd2_data  = [['Unidad DVD2:', 'Marca:', valores_celdas[0], 'Modelo:', valores_celdas[1], 
                                        'Capacidad:', valores_celdas[2], 'N° Serie:', valores_celdas[3]]]
                
    

                dvd2_table = Table(dvd2_data, colWidths=[0.9*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch, 0.8*inch, 1.5*inch ])
                dvd2_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('PADDING', (0, 0), (-1, -1), 3),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ]))
                elements.append(dvd2_table)
                
                dvd2_table.hAlign = 'LEFT'
                elements.append(Spacer(1, 10))

            # Diagnóstico y recomendaciones
            
           
          

            elements.append(Paragraph("DIAGNÓSTICO Y RECOMENDACIONES:", title_style))
            elements.append(Spacer(1, 10))

            sections = [
                ("Descripción de la falla", get_cell_value(105)),
                ("Diagnóstico", get_cell_value(106)),
                ("Reparación", get_cell_value(107))
            ]

            notas = get_cell_value(108)
            if notas != "N/A":  # Filtra la sección si el valor es "N/A"
                sections.append(("Notas", notas))

            for title, content in sections:
                elements.append(Paragraph(title, styles['Heading2']))
                
                # Dividir el contenido en párrafos si es necesario
                paragraphs = content.split("\n")  # Suponiendo que los párrafos están separados por doble salto de línea
                for paragraph in paragraphs:
                    elements.append(Paragraph(paragraph, normal_style))
                    elements.append(Spacer(1, 10))  # Espaciado entre párrafos

                elements.append(Spacer(1, 20))  # Espaciado entre secciones

                # Agregar salto de página si el contenido es muy largo
                if len(content) > 3000:  # Ajusta este valor según el espacio disponible
                    elements.append(PageBreak())

            # Generar el PDF
            doc.build(elements)

            QMessageBox.information(self, "Éxito", f"Archivo PDF generado exitosamente.\nRuta: {file_name}")

        except Exception as e:
            print(f"Error al generar el archivo PDF: {e}")
            QMessageBox.warning(self, "Error", "Ocurrió un error al generar el archivo PDF.")

    def generate_excel(self):
        try:
            # Obtener la fila seleccionada
            selected_items = self.table.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "Selección requerida", "Por favor, seleccione un registro de la tabla.")
                return

            row = selected_items[0].row()

            # Obtener datos de la fila seleccionada
            def get_cell_value(col):
                item = self.table.item(row, col)
                return item.text() if item else "N/A"

            # Crear el archivo Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hoja de Vida del Equipo"
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10

            # Información del cliente
            ws['A1'] = "DATOS DEL USUARIO:"
            ws.merge_cells('A1:B1')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            ws['A2'] = "Cliente:"
            ws['B2'] = get_cell_value(1)  # cliente
            ws['A3'] = "Cedula / NIT:"
            ws['B3'] = get_cell_value(2)  # cedula_nit
            ws['A4'] = "Correo:"
            ws['B4'] = get_cell_value(3)  # correo
            ws['A5'] = "Teléfono:"
            ws['B5'] = get_cell_value(4)  # telefono
            ws['A6'] = "Fecha Ingreso:"
            ws['B6'] = get_cell_value(6)  # fecha_ingreso
            ws['A7'] = "Fecha Entrega:"
            ws['B7'] = get_cell_value(7)  # fecha_entrega

            # Título de datos del equipo
            ws['A9'] = "DATOS DEL EQUIPO:"
            ws.merge_cells('A9:B9')
            ws['A9'].alignment = Alignment(horizontal='center', vertical='center')

            # Información del equipo
            basic_data = [
                ['Tipo:', get_cell_value(8)],
                ['Marca:', get_cell_value(9)],
                ['Modelo:', get_cell_value(10)],
                ['N° Serie:', get_cell_value(11)],
            ]

            for i, (label, value) in enumerate(basic_data, start=10):
                ws[f'A{i}'] = label
                ws[f'B{i}'] = value

            current_row = 14  # Iniciar en la fila 14

           # Obtener los valores de la Main Board
            marca_main = get_cell_value(12)  # Marca de la Main Board
            modelo_main = get_cell_value(13)  # Modelo de la Main Board
            serie_main = get_cell_value(14)  # N° Serie de la Main Board

            # Verificar si todos los valores son "N/A"
            if not all(valor == "N/A" for valor in [marca_main, modelo_main, serie_main]):
                # Si al menos un valor no es "N/A", agregar la fila a la hoja de cálculo
                ws['A' + str(current_row)] = "Main Board:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_main
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_main
                ws['F' + str(current_row)] = "N° Serie:"
                ws['G' + str(current_row)] = serie_main
                current_row += 1
                
                
                
            # WiFi
            marca_wifi=get_cell_value(15)
            modelo_wifi=get_cell_value(16)
            serie_wifi=get_cell_value(17)
            
            if not all(valor == "N/A" for valor in [marca_wifi, modelo_wifi, serie_wifi]):
                ws['A' + str(current_row)] = "WiFi:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_wifi
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_wifi
                ws['F' + str(current_row)] = "N° Serie:"
                ws['G' + str(current_row)] = serie_wifi
                current_row += 1

            # Teclado
            marca_teclado=get_cell_value(18)
            modelo_teclado=get_cell_value(19)
            serie_teclado=get_cell_value(20)
            
            if not all(valor == "N/A" for valor in [marca_teclado, modelo_teclado, serie_teclado]):
            
                ws['A' + str(current_row)] = "Teclado:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_teclado
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_teclado
                ws['F' + str(current_row)] = "N° Serie:"
                ws['G' + str(current_row)] = serie_teclado
                current_row += 1

            # CPU1
            
            marca_CPU1=get_cell_value(21)
            modelo_CPU1=get_cell_value(22)
            serie_CPU1=get_cell_value(23)
            
            if not all(valor == "N/A" for valor in [marca_CPU1, modelo_CPU1, serie_CPU1]):
            
                ws['A' + str(current_row)] = "CPU1:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_CPU1
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_CPU1
                ws['F' + str(current_row)] = "N° Serie:"
                ws['G' + str(current_row)] = serie_CPU1
                current_row += 1

            # CPU2
            marca_CPU2=get_cell_value(24)
            modelo_CPU2=get_cell_value(25)
            serie_CPU2=get_cell_value(26)
            
            if not all(valor == "N/A" for valor in [marca_CPU2, modelo_CPU2, serie_CPU2]):
                
                ws['A' + str(current_row)] = "CPU2:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_CPU2
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_CPU2
                ws['F' + str(current_row)] = "N° Serie:"
                ws['G' + str(current_row)] = serie_CPU2
                current_row += 1

            # PSU
            marca_PSU=get_cell_value(27)
            modelo_PSU=get_cell_value(28)
            capacidad_PSU=get_cell_value(29)
            
            if not all(valor == "N/A" for valor in [marca_PSU, modelo_PSU, capacidad_PSU]):
                ws['A' + str(current_row)] = "PSU:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_PSU
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_PSU
                ws['F' + str(current_row)] = "Capacidad:"
                ws['G' + str(current_row)] =  capacidad_PSU
                current_row += 1

            # Pantalla
            marca_pantalla=get_cell_value(30)
            modelo_pantalla=get_cell_value(31)
            capacidad_pantalla=get_cell_value(32)
            serie_pantalla=get_cell_value(33)
            
            
            if not all(valor == "N/A" for valor in [ marca_pantalla, modelo_pantalla,  capacidad_pantalla,serie_pantalla]):
                ws['A' + str(current_row)] = "Pantalla:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] =  marca_pantalla
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_pantalla
                ws['F' + str(current_row)] = "Capacidad:"
                ws['G' + str(current_row)] = capacidad_pantalla
                ws['H' + str(current_row)] = "N° Serie:"
                ws['I' + str(current_row)] =  serie_pantalla
            
                current_row += 1

            # Batería
            marca_bateria=get_cell_value(34)
            modelo_bateria=get_cell_value(35)
            capacidad_bateria=get_cell_value(37)
            serie_bateria=get_cell_value(36)
            
            
            if not all(valor == "N/A" for valor in [ marca_bateria, modelo_bateria,  capacidad_bateria,serie_bateria]):
                ws['A' + str(current_row)] = "Batería:"
                ws['B' + str(current_row)] = "Marca:"
                ws['C' + str(current_row)] = marca_bateria
                ws['D' + str(current_row)] = "Modelo:"
                ws['E' + str(current_row)] = modelo_bateria
                ws['F' + str(current_row)] = "Capacidad:"
                ws['G' + str(current_row)] =  capacidad_bateria
                ws['H' + str(current_row)] = "N° Serie:"
                ws['I' + str(current_row)] = serie_bateria
                current_row += 1

            # RAM (4 módulos)
            for i in range(1, 5):
                base_idx = 38 + (i-1)*4  # Índice base para RAM
                marca = get_cell_value(base_idx)
                modelo = get_cell_value(base_idx + 1)
                capacidad = get_cell_value(base_idx + 2)
                serie = get_cell_value(base_idx + 3)
            
            # Verificar si todos los valores son "N/A"
                if not all(valor == "N/A" for valor in [marca, modelo, capacidad, serie]):
                    ws['A' + str(current_row)] = f"RAM {i}:"
                    ws['B' + str(current_row)] = "Marca:"
                    ws['C' + str(current_row)] = marca
                    ws['D' + str(current_row)] = "Modelo:"
                    ws['E' + str(current_row)] = modelo
                    ws['F' + str(current_row)] = "Capacidad:"
                    ws['G' + str(current_row)] = capacidad
                    ws['H' + str(current_row)] = "N° Serie:"
                    ws['I' + str(current_row)] = serie
                    current_row += 1
                    
               # Unidades de almacenamiento (8 unidades)
     
            for i in range(1, 9):
                base_idx = 54 + (i-1)*4  # Índice base para RAM
                marca = get_cell_value(base_idx)
                modelo = get_cell_value(base_idx + 1)
                capacidad = get_cell_value(base_idx + 2)
                serie = get_cell_value(base_idx + 3)
            
            # Verificar si todos los valores son "N/A"
                if not all(valor == "N/A" for valor in [marca, modelo, capacidad, serie]): 
           
                    ws['A' + str(current_row)] = f"Unidad Almacenamiento {i}:"
                    ws['B' + str(current_row)] = "Marca:"
                    ws['C' + str(current_row)] = get_cell_value(base_idx)
                    
                    ws['D' + str(current_row)] = "Modelo:"
                    ws['E' + str(current_row)] = get_cell_value(base_idx + 1)
                    
                    ws['F' + str(current_row)] = "Capacidad:"
                    ws['G' + str(current_row)] = get_cell_value(base_idx + 2)
                    
                    ws['H' + str(current_row)] = "N° Serie:"
                    ws['I' + str(current_row)] = get_cell_value(base_idx + 3)
                    
                    
                    current_row += 1
            for i in range(1, 3):
                base_idx = 86 + (i-1)*4  
                marca = get_cell_value(base_idx)
                modelo = get_cell_value(base_idx + 1)
                capacidad = get_cell_value(base_idx + 2)
                serie = get_cell_value(base_idx + 3)
            
            # Verificar si todos los valores son "N/A"
                if not all(valor == "N/A" for valor in [marca, modelo, capacidad, serie]): 
                    ws['A' + str(current_row)] = f"GPU {i}:"
                    ws['B' + str(current_row)] = "Marca:"
                    ws['C' + str(current_row)] = get_cell_value(base_idx)
                    
                    ws['D' + str(current_row)] = "Modelo:"
                    ws['E' + str(current_row)] = get_cell_value(base_idx + 1)
                    
                    
                    ws['F' + str(current_row)] = "Capacidad:"
                    ws['G' + str(current_row)] = get_cell_value(base_idx + 2)  #Esta invertido en la base de datos
                    
                    
                    ws['H' + str(current_row)] = "N° Serie:"
                    ws['I' + str(current_row)] = get_cell_value(base_idx + 3)
                    current_row += 1
            
            for i in range(1, 3):
                base_idx = 94 + (i-1)*4  
                marca = get_cell_value(base_idx)
                modelo = get_cell_value(base_idx + 1)
                capacidad = get_cell_value(base_idx + 2)
                serie = get_cell_value(base_idx + 3)
            
            # Verificar si todos los valores son "N/A"
                if not all(valor == "N/A" for valor in [marca, modelo, capacidad, serie]):         
           
                    ws['A' + str(current_row)] = f"UNIDAD DVD {i}:"
                    ws['B' + str(current_row)] = "Marca:"
                    ws['C' + str(current_row)] = get_cell_value(base_idx)
                    
                    ws['D' + str(current_row)] = "Modelo:"
                    ws['E' + str(current_row)] = get_cell_value(base_idx + 1)
                    
                    
                    ws['F' + str(current_row)] = "Capacidad:"
                    ws['G' + str(current_row)] = get_cell_value(base_idx + 2)
                    
                    
                    ws['H' + str(current_row)] = "N° Serie:"
                    ws['I' + str(current_row)] = get_cell_value(base_idx + 3)
                    current_row += 1
                    
                        # Título de fallas 
                ws['A40'] = "DIAGNÓSTICO Y RECOMENDACIONES:"
                ws.merge_cells('A40:D40')
                ws['A40'].alignment = Alignment(horizontal='center', vertical='center')
                
                
                ws['A42'] = "Descripción de Falla"
                ws.merge_cells('A42:D42')
                ws['A42'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws['A43'] = get_cell_value(105)
                ws.merge_cells('A43:D43')
                ws['A43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                ws['A45'] = "Diagnostico"
                ws.merge_cells('A45:D45')
                ws['A45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws['A46'] = get_cell_value(106)
                ws.merge_cells('A46:D46')
                ws['A46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                ws['A48'] = "Reparación"
                ws.merge_cells('A48:D48')
                ws['A48'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                ws['A49'] = get_cell_value(107)
                ws.merge_cells('A49:D49')
                ws['A49'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                
                notas= get_cell_value(108)

                if notas != "N/A":
                    ws['A51'] = "Notas"
                    ws.merge_cells('A51:D51')
                    ws['A51'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                    ws['A52'] = notas
                    ws.merge_cells('A52:D52')
                    ws['A52'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                ws.row_dimensions[43].height = 100
                ws.row_dimensions[46].height = 100
                ws.row_dimensions[49].height = 100
                ws.row_dimensions[52].height = 100
                

            # Seleccionar ubicación y guardar el archivo
            options = QFileDialog.Options()
            file_path, _ = QFileDialog.getSaveFileName(self, "Guardar Archivo", ".xlsx", "Archivos Excel (*.xlsx)", options=options)
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, "Éxito", "El archivo se ha guardado correctamente.")
            else:
                QMessageBox.warning(self, "Cancelado", "No se guardó el archivo.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Ocurrió un error al generar el archivo: {str(e)}")

    def salir(self):
    
        QApplication.quit()  # Cierra la aplicación



if __name__ == "__main__":
    app = QApplication(sys.argv)
    window= Tab4()
    window.show()
    sys.exit(app.exec_())
